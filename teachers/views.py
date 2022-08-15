from asyncore import write
from urllib import response
from django.http import HttpResponse
from django.shortcuts import render , redirect
from django.contrib.auth import logout
from yaml import load
from home import views
from datetime import datetime as dt
from django.contrib.auth.decorators import login_required
import uuid , re
import qrcode
from PIL import Image , ImageDraw , ImageFont
from .models import TeacherData
import csv
import xlwt
from django.contrib.auth.hashers import check_password
#---------------------------------------
# Firebase Module Block
#---------------------------------------
import pyrebase
import os

from verifID.settings import BASE_DIR, HOME_DIR
from . import credentials
#Firebase Connection
def FirebaseConnection():
    firebaseConfigdefault = {
                        'apiKey': credentials.FIREBASE_API_KEY,
                        'authDomain': credentials.FIREBASE_AUTH_DOMAIN,
                        'projectId': credentials.FIREBASE_PROJECT_ID,
                        'databaseURL': credentials.FIREBASE_DATABASE_URL,
                        'storageBucket': credentials.FIREBASE_STORAGE_BUCKET,
                        'messagingSenderId': credentials.FIREBASE_SENDER_ID,
                        'appId': credentials.FIREBASE_APP_ID,
                        'measurementId': credentials.FIREBASE_MEASUREMENT_ID,
    }
    firebase = pyrebase.initialize_app(firebaseConfigdefault)
    return firebase
#Firebase Storage Initialize
def fireStore():
    myFireStore = FirebaseConnection().storage()
    return myFireStore
# Firebase Real Time Database Initialize
def fireData():
    myFireDB = FirebaseConnection().database()
    return myFireDB
#Individual Routine Access
def get_routine(teacherName):
    destinationPath = teacherName+"/routine.jpg"
    fireStoreObject = fireStore()
    imgurl =  fireStoreObject.child(destinationPath).get_url(None)
    return imgurl
#Download Routine 
import requests
def get_routineDownloaded(request):
    fname = request.user.first_name
    lname = request.user.last_name
    teacherName = fname + lname
    downPath = teacherName+"/routine.jpg"
    fireStoreObject = fireStore()
    downloadURL = fireStoreObject.child(downPath).get_url(None)
    response = HttpResponse(content_type = "image/png")
    response['Content-Disposition'] = "attachment;filename=image.png"
    response.write(requests.get(downloadURL).content)
    return response
#Download ID
def get_idDownloaded(request):
    fname = request.user.first_name
    lname = request.user.last_name
    teacherName = fname + lname
    downPath = teacherName+"/idCard.png"
    fireStoreObject = fireStore()
    downloadURL = fireStoreObject.child(downPath).get_url(None)
    response = HttpResponse(content_type = "image/png")
    response['Content-Disposition'] = "attachment;filename=idCard.png"
    response.write(requests.get(downloadURL).content)
    return response
#--------------------------------
# End
#--------------------------------
# MongoDB Atlas Module Block
#--------------------------------
import pymongo
from pymongo import MongoClient
from . import credentials
def mongoAttendanceDB(timeStamp,subjectCode,attendanceList):
    # For Successful Insertion
    insertFlag = True
    # cluster URL
    clusterURL = "mongodb+srv://"+credentials.MONGODB_USERNAME+":"+credentials.MONGODB_PASSWORD+"@attendanccedb.vkkyk.mongodb.net/?retryWrites=true&w=majority"
    # setup Connection with cluster
    myCluster = MongoClient(clusterURL,tls=True,tlsAllowInvalidCertificates=True)
    # setup connection with database
    myDB = myCluster["attendanceMCKVIE"]
    # setup connnection with collection
    myCollection = myDB[subjectCode]
    myData = {
        'timeStamp':timeStamp,
        'hasData':True,
        'AttendanceList':attendanceList,
    }
    checkDuplicate = myCollection.find_one({'timeStamp':timeStamp})
    if checkDuplicate is not None:
        checkDuplicateObject = checkDuplicate['hasData']
    else:
        checkDuplicateObject = False
    if checkDuplicateObject is True:
        insertFlag = False
        pass
    else:
        myCollection.insert_one(myData)
    return insertFlag

def archive(request):
    if request.method == "POST":
        fileReset()
        subject_code = request.POST['subjectcode']
        response = HttpResponse(content_type = "application/ms-excel")
        response['Content-Disposition'] = "attachment;filename=attendance"+ subject_code +".xlsx"
        path = r'attendanceArchive/attendance'+subject_code+'.xlsx'
        downlaodPath = fireStore().child(path).get_url(None)
        response.write(requests.get(downlaodPath).content)
        return response
    return render(request,'teachers/getArchive.html')

def getClassStrength(stream):
    # cluster URL
    clusterURL = "mongodb+srv://"+credentials.MONGODB_USERNAME+":"+credentials.MONGODB_PASSWORD+"@attendanccedb.vkkyk.mongodb.net/?retryWrites=true&w=majority"
    # setup connection with cluster
    myCluster = MongoClient(clusterURL,tls=True,tlsAllowInvalidCertificates=True)
    # setup connection with database
    myDB = myCluster["studentData"]
    # setup connnection with collection
    myCollection = myDB[stream]
    data = myCollection.find_one({'batch':'2019-23','currentSem':6})
    totalStrength = data['totalStrength']
    return totalStrength

def setAncLog(collectionName,filePath,subjectCode,about,dateOfSubmit):
    # cluster URL
    clusterURL = "mongodb+srv://"+credentials.MONGODB_USERNAME+":"+credentials.MONGODB_PASSWORD+"@attendanccedb.vkkyk.mongodb.net/?retryWrites=true&w=majority"
    # setup connection with cluster
    myCluster = MongoClient(clusterURL,tls=True,tlsAllowInvalidCertificates=True)
    # setup connection with database
    myDB = myCluster["announcementData"]
    # setup connnection with collection
    myCollection = myDB[collectionName]
    timeStamp = dt.now().strftime('%d-%m-%y')
    data = {
        'subjectCode':subjectCode,
        'about':about,
        'datePosted': timeStamp,
        'dateOfSubmit':dateOfSubmit,
        'filePath':filePath,
    }
    myCollection.insert_one(data)

def sendMessage(request):
    if request.method == 'POST':
        message = request.POST['message']
        # cluster URL
        clusterURL = "mongodb+srv://"+credentials.MONGODB_USERNAME+":"+credentials.MONGODB_PASSWORD+"@attendanccedb.vkkyk.mongodb.net/?retryWrites=true&w=majority"
        # setup connection with cluster
        myCluster = MongoClient(clusterURL,tls=True,tlsAllowInvalidCertificates=True)
        # setup connection with database
        myDB = myCluster["adminMessage"]
        # setup connnection with collection
        myCollection = myDB["teachers"]
        timeStamp = dt.now().strftime('%d-%m-%y')
        data = {
            'Date':timeStamp,
            'Teacher':request.user.first_name + " " + request.user.last_name,
            'Message':message
        }
        myCollection.insert_one(data)
        return redirect('adminMessage')
    return render(request,'teachers/adminMessage.html')
#---------------------------------
# End
#---------------------------------


# My views here.
@login_required(login_url='home')
# @allowed_users(allowed_roles=['teachers'])
def dashboard(request):
    group = request.user.groups.all()[0].name
    if group == 'teachers':
        pass
    else:
        return render(request,'teachers/maintainence.html',context = {'message':'Not Authorized'})
    fname = request.user.first_name
    lname = request.user.last_name
    fullname = fname + " " + lname
    # useremail = request.user.email
    userObj = TeacherData.objects.get(email = 'manosijindia@gmail.com')
    designation = userObj.designation
    department = userObj.department
    cid = userObj.collegeid
    profileUrl = fireStore().child(fname+lname+'/profilepicture.jpg').get_url(None)
    # print(profileUrl)
    context = {
        'fullname':fullname,
        'designation':designation,
        'department':department,
        'cid':cid,
        'profilepicture':profileUrl,
    }
    return render(request,'teachers/dashboard.html',context)    

def routineTeacher(request):
    fname = request.user.first_name
    lname = request.user.last_name
    if request.method == "POST":
        if request.POST.get('download_routine'):
            return redirect('downloadRoutine')
        else:
            return render(request,'teachers/maintainence.html',context = {'message':'Feature Coming Soon'})
    img_url = get_routine(fname+lname)
    context = {
        'url' :img_url
    }
    #os.chdir(HOME_DIR)
    return render(request,'teachers/routineDash.html',context)

def announcement(request):
    if request.method == "POST":
        announcementText = request.POST['announcement']
        subjectCode = request.POST['subjectcodeanc']
        timeStamp = dt.now().strftime('%d/%m/%y')
        teacherName = request.user.get_full_name()
        data = {
            'date':timeStamp,
            'teacher':teacherName,
            'text':announcementText,
            'seen':False,
            'subjectCode':subjectCode,
        }
        myFireDBObject = fireData()
        myFireDBObject.push(data)
    return render(request,'teachers/postAnnouncement.html')

def attendance(request):
    if request.method == "POST":
        subjectCode = request.POST["subjectCode"]
        attendanceList = request.POST["attendanceList"]
        try:
            attendanceList = list(map(int,attendanceList.split(',')))
        except:
            attendanceList = list(map(int,attendanceList.split(' ')))
        logBook = list()
        logBook.append('null')
        for i in range(1,getClassStrength("CSE")+1):
            if i in attendanceList:
                logBook.append('P')
            else:
                logBook.append('A')
        timeStamp = dt.now().strftime('%d/%m/%y')
        returnValue = mongoAttendanceDB(timeStamp,subjectCode,logBook)
        if returnValue is False:
            return render(request,'teachers/maintainence.html',context = {'message':'Data Record Present'})
        else:
            writeFile(timeStamp,subjectCode,logBook)
    return render(request,'teachers/setAttendance.html')

def myid(request):
    teacherDataObject = TeacherData.objects.all().get(email = request.user.email)
    # # get device id
    # macAddress = ':'.join(re.findall('..', '%012x' % uuid.getnode()))
    # # uniqueid
    # uniqueId = teacherDataObject.deviceid
    # if macAddress != uniqueId:
    #     return HttpResponse("Device not Registered")
    # make qr
    qr_img = qrcode.make(teacherDataObject.collegeid) 
    #os.chdir(os.path.join(HOME_DIR+"\\etc"))
    qr_img.save("etc/qr.jpg")
    # make idcard
    template = Image.open("etc/template.png")
    draw = ImageDraw.Draw(template)
    font = ImageFont.truetype("etc/Ubuntu-Regular.ttf",size=26)
    draw.text((366,245),dt.now().strftime('%d/%m/%Y'),fill = "black",font=font)
    nameOfUser = teacherDataObject.firstname + " " + teacherDataObject.lastname
    draw.text((376,115),nameOfUser,fill = "black",font=font)
    draw.text((473,160),teacherDataObject.designation,fill = "black",font=font)
    draw.text((371,210),teacherDataObject.department,fill = "black",font=ImageFont.truetype("etc/Ubuntu-Regular.ttf",size=20))
    pic = Image.open('etc/qr.jpg').resize((196,218),Image.ANTIALIAS)
    template.paste(pic,(61,116,257,334))
    template.save("etc/idCard.png")
    fireStoreObject = fireStore()
    path = request.user.first_name + request.user.last_name + "/idCard.png"
    # path = request.user.first_name + request.user.last_name
    fireStoreObject.child(path).put("etc/idCard.png")
    return redirect('downloadid')

def uploadAssignment(request):
    if request.method == "POST":
        fileName = dt.now().strftime("%d%m%Y%H%M%S") + "assignment.jpeg"
        fireStoreObject = fireStore()
        path = request.user.first_name + request.user.last_name + '/upload/' + fileName
        fireStoreObject.child(path).put(request.FILES['asgimg'])
        filePath = fireStoreObject.child(path).get_url(None)
        description = request.POST['aboutannc']
        subjectCode = request.POST['subjectcode']
        dateOfSubmit = request.POST['dos']
        setAncLog("collection01",filePath,subjectCode,description,dateOfSubmit)
    return render(request,'teachers/sendAssignments.html')
#includes
from django.contrib.auth.models import User

def updatePassword(request):
    if request.method == "POST":
        oldPass = request.POST['oldpassword']
        newPass = request.POST['newpassword']
        confNewPass = request.POST['confpassword']
        userPassword = request.user.password
        checkForAuth = check_password(oldPass,userPassword)
        if checkForAuth is False or newPass != confNewPass:
            return HttpResponse("Auth Denied")
        uname = request.user.username
        uobj = User.objects.get(username=uname)
        uobj.set_password(newPass)
        uobj.save()
        return redirect('updatePassword')   
    return render(request,'teachers/updatePassword.html')

def updateProfile(request):
    if request.method == "POST":
        # newEmail = request.POST['email']
        newPhone = request.POST['phoneno']
        newAddr = request.POST['address']
        userObj = TeacherData.objects.get(email = request.user.email)
        # userObj.email = newEmail
        userObj.phoneno = newPhone
        userObj.address = newAddr
        userObj.save()
        return redirect('home')
    # email = request.user.email
    phone = request.user.teacherdata.phoneno
    address = request.user.teacherdata.address
    context = {
        # 'email':email,
        'phone':phone,
        'address':address,
    }
    return render(request,'teachers/updateProfile.html',context)        

from openpyxl import load_workbook
def writeFile(timeStamp,subjectCode,logBook):
    filePath = r'etc/attendance'+subjectCode+'.xlsx'
    file = load_workbook(filePath)
    sheet = file.active
    cols = sheet.max_column
    sheet.cell(row=1,column=cols+1).value = timeStamp
    totalStudents = getClassStrength('CSE')
    j = 1
    for i in range(2,totalStudents+1):
        sheet.cell(row=i,column=cols+1).value = logBook[j]
        j = j + 1
    file.save(filePath)
    storage = fireStore()
    fileName = r'attendanceArchive/attendance'+subjectCode+'.xlsx'
    storage.child(fileName).put(filePath)

def fileReset():
    path = r'attendanceArchive/attendancePCC-CS601'+'.xlsx'
    print(BASE_DIR)
    os.chdir(os.path.join(BASE_DIR,'etc'))
    fireStore().child(path).download(path='',filename='attendancePCC-CS601.xlsx')
    os.chdir(BASE_DIR)
###################################################################
#TESTING CODE
###################################################################
# from django.contrib.auth.hashers import check_password
def test(request):
    path = r'attendanceArchive/attendancePCC-CS601'+'.xlsx'
    print(BASE_DIR)
    os.chdir(os.path.join(BASE_DIR,'etc'))
    fireStore().child(path).download(path='',filename='attendancePCC-CS601.xlsx')
    os.chdir(BASE_DIR)
    return HttpResponse("TEST")
###################################################################
###################################################################